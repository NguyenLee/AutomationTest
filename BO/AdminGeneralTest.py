import sys
SertaFolder = 'D:\AutomationTest\Scripts\SertaTA'
if SertaFolder not in sys.path:
    sys.path.append(SertaFolder)
from Util import Util
import unittest, time
import os, logging
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import openpyxl as pyxl

from AdminUrls import AdminUrls
from AdminLoginTest import AdminLoginTest

class AdminGeneralTest(unittest.TestCase):
    def setUp(self):
        self.driver = webdriver.Firefox()
        self.driver.get(AdminUrls.LOGIN_URL)
        AdminLoginTest.testLoginCorrectAccount(self)

    def testMenuEleLinks(self, eleNo):
        itemMenuEles = self.driver.find_elements_by_xpath("//div[@class='sidebar-left-content']/ul[@class='sidebar-menu']/li[" + str(eleNo) + "]/ul/li/url[@class='sub-nav']/li")
        links = []
        for itemEle in itemMenuEles:
            menuLink = itemEle.find_element_by_tag_name('a').get_attribute('href')
            #print(itemEle.get_attribute('text'), ': ', menuLink)
            links.append(menuLink)

        for link in links:
            try :
                print('*** Link: ', link)
                print('Page title: ', self.driver.title)
                self.driver.get(link)
                time.sleep(15)
            except Warning as w:
                print(w)
            except Exception as e:
                logging.exception("message")

    def testAllMenuItemLinks(self):
        self.driver.get(AdminUrls.BASE_URL)
        WebDriverWait(self.driver, 15).until(EC.presence_of_element_located((By.ID, 'qa-sidebar-menu')))
        ulEle = self.driver.find_element_by_class_name('sidebar-menu')
        liEles = ulEle.find_elements_by_xpath("//ul[@id='qa-sidebar-menu']/li[@class='qa-group-function']")

        # Write menu items on Excel file
        wb = pyxl.Workbook()
        ws = wb.active
        ws.title = 'Admin'
        links = []
        ws['A1'] = 'Main menu'
        ws['B1'] = 'Menu item'
        ws['C1'] = 'Link'
        ws['D1'] = 'Result'
        row = 1
        for liEle in liEles :
            row = row + 1
            groupFunc = liEle.find_element_by_xpath(".//a[@class='accordion-toggle']/span[@class='sidebar-title']").get_attribute('innerHTML')
            print('*** ', groupFunc)
            ws['A' + str(row)] = groupFunc
            funcEles = liEle.find_elements_by_xpath(".//ul/li/a")
            for i in range(0, len(funcEles)) :
                ele = funcEles[i]
                # Item name
                func = ele.get_attribute('text')
                print(' - ', func)
                ws['B' + str(row)] = func
                ws['C' + str(row)] = ele.get_attribute('href')
                if (i < len(funcEles) - 1) :
                    row = row + 1

                # List page to go through
                menuLink = ele.get_attribute('href')
                links.append(menuLink)
        excelFileName = 'D:\AutomationTest\Scripts\SertaTA\DataTest\AdminCheckList.xlsx'
        if (os.path.isfile(excelFileName)) :
            os.remove(excelFileName)
        wb.save(excelFileName)

        # Go through all pages to check exception
        LOG_FILENAME = 'D:\AutomationTest\Scripts\SertaTA\AdminLogging.txt'
        logging.basicConfig(filename=LOG_FILENAME,
                            level=logging.ERROR)
        wb = pyxl.load_workbook(excelFileName)
        ws = wb.get_sheet_by_name('Admin')
        row = 2
        print('***** GO THROUGH PAGES: *****')
        for link in links:
            try :
                print('* ', link)
                self.driver.get(link)
                time.sleep(15)
                result = 'OK'
                if (self.driver.title.lower().find('php warning') > -1 or self.driver.title.lower().find('exception') > -1 or self.driver.title.lower().find('error') > -1) :
                    msg = self.driver.find_element_by_class_name('message').get_attribute('innerHTML')
                    logging.error(link + ': ' + msg)
                    result = msg
                row = row + 1
            except Warning as w:
                logging.warning('Warning: ' + w)
            except Exception as e:
                logging.exception('Exception: ' + e)
            print(result)
            ws['D' + str(row)] = result
        wb.save(excelFileName)

    def testWarning(self):
        LOG_FILENAME = 'D:\AutomationTest\Scripts\SertaTA\AdminLogging.txt'
        if (os.path.isfile(LOG_FILENAME)):
            os.remove(LOG_FILENAME)
        logging.basicConfig(filename=LOG_FILENAME,
                            level=logging.WARNING)
        try:
            link = "https://dev2.serta.com/admin/BazaarVoice/admin/BazaarVoiceProductSettings"
            self.driver.get(link)
            time.sleep(15)
            if (self.driver.title.lower().find('php warning') > -1 or self.driver.title.lower().find(
                    'exception') > -1 or self.driver.title.lower().find('error') > -1):
                msg = self.driver.find_element_by_class_name('message').get_attribute('innerHTML')
                print(link, ': ', self.driver.title, ':', msg)
                logging.warning(link + ': ' + msg)
        except Warning as w:
            logging.warning(w)

    def tearDown(self):
        self.driver.close()

if __name__ == '__main__':
    # Testcases are ordered by alphabet
    suite = unittest.TestSuite()
    suite.addTest(AdminGeneralTest('testAllMenuItemLinks'))
    #suite.addTest(AdminGeneralTest('testWarning'))
    #suite.addTest(AdminGeneralTest('testReadLogFile'))
    unittest.TextTestRunner(verbosity=2).run(suite)